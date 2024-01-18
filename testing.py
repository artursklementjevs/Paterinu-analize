import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

Wb_test = Workbook()
Wb_test.save("analize.xlsx")
wb = openpyxl.Workbook()


sheet = wb.active

rimi_nauda = 0
expo_nauda = 0
cita_nauda = 0
avota_nauda = 0

i = 1

rinda = 2


df = pd.read_excel('example.xls')
pieraksts = sheet.cell(row = 1, column = 1)
pieraksts.value = "Avota nosaukums"

pieraksts = sheet.cell(row = 1, column = 2)
pieraksts.value = "Patēriņa summa"


########################### DATU IEVADE ##############################################


budzets = input("Sveicināts! Kāds ir šī mēneša budžets: ")
budzets = float(budzets)
skaits = input("Cik patēriņu avotus vēlies apskatīt?: ")
skaits = float(skaits)


########################### DATU KALKULATORS ########################################

while (i<=skaits):
    avots = input("Ievadi avota nosaukumu: ")
    pieraksts = sheet.cell(row = rinda, column = 1)
    pieraksts.value = avots

    kartosana = df['Unnamed: 2'].str.contains(avots, case=False, na=False)
    avota_df = df[kartosana]
    for summa in avota_df['Unnamed: 5']:
        if summa<0 :
            avota_nauda = avota_nauda + summa

    pieraksts = sheet.cell(row=rinda, column=2)
    pieraksts.value = avota_nauda
    cita_nauda = cita_nauda - avota_nauda
    print(avots + " patēriņš : " + str(round(avota_nauda, 2)) + "€")


    avota_nauda = 0
    i = i+1
    rinda = rinda + 1

kartosana = df['Unnamed: 2'].str.contains("RIMI", case=False, na=False)
rimi_df = df[kartosana]
kartosana = df['Unnamed: 2'].str.contains("RTU-BT1", case=False, na=False)
expo_df = df[kartosana]
kartosana = df['Unnamed: 4'].str.contains("Debeta apgrozījums", case=False, na=False)
parejais_df = df[kartosana]

for summa in rimi_df['Unnamed: 5']:
    rimi_nauda = rimi_nauda + summa

for summa in expo_df['Unnamed: 5']:
    expo_nauda = expo_nauda + summa

for summa in parejais_df['Unnamed: 5']:
    cita_nauda = cita_nauda + summa
    kopeja_nauda = summa

pieraksts = sheet.cell(row=rinda, column=1)
pieraksts.value = "RIMI"
pieraksts = sheet.cell(row=rinda, column=2)
pieraksts.value = round(rimi_nauda, 2)
rinda = rinda + 1
pieraksts = sheet.cell(row=rinda, column=1)
pieraksts.value = "EXPO"
pieraksts = sheet.cell(row=rinda, column=2)
pieraksts.value = round(expo_nauda, 2)
rinda = rinda + 1
cita_nauda = cita_nauda - rimi_nauda - expo_nauda
pieraksts = sheet.cell(row=rinda, column=1)
pieraksts.value = "Pārējais"
pieraksts = sheet.cell(row=rinda, column=2)
pieraksts.value = round(cita_nauda, 2)

wb.save("analize.xlsx")
########################### DATU IZVADE ########################################




print("\nPatēriņš Rimi: " + str(round(rimi_nauda, 2))+"€")
print("Patēriņš EXPO: " + str(round(expo_nauda, 2))+"€")

print("Pārējais patēriņs: " + str(round(cita_nauda, 2))+"€")


print("\nKopējais patēriņs: " + str(round(kopeja_nauda, 2))+"€")

print("Mēneša atlikums:" + str(budzets + round(kopeja_nauda, 2))+"€")




########################### DATU IEVADE GRAFIKĀ ##################################


wb = openpyxl.load_workbook('analize.xlsx')
ws = wb.active
pica = PieChart()


avoti = Reference(ws, min_col=1, min_row=2, max_row=rinda)
dati = Reference(ws, min_col=2, min_row=2, max_row=rinda)

pica.add_data(dati)
pica.set_categories(avoti)
pica.title = ws.cell(1, 1).value
pica.title = "Mēneša patēriņš"
ws.add_chart(pica, 'D1')
wb.save('analize.xlsx')
wb.close()